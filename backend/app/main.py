from fastapi import FastAPI, Depends, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
import io
from openpyxl import Workbook
import json
from .db import get_db
from . import crud, schemas
from fastapi.middleware.cors import CORSMiddleware
from .db import engine
from . import models

app = FastAPI(title="Fajr Global Clinic System (4-table)")
models.Base.metadata.create_all(bind=engine)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://fajerg.netlify.app",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://localhost:5500",
        "http://127.0.0.1:5500",
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="app/static"), name="static")

@app.get("/")
def index(): return FileResponse("app/static/index.html")

@app.post("/api/patients", response_model=schemas.PatientOut)
def create_patient(payload: schemas.PatientCreate, db: Session = Depends(get_db)):
    p = crud.create_patient(db, payload)
    return schemas.PatientOut(id=p.id, full_name=p.full_name, date_of_birth=p.date_of_birth, sex=p.sex, no_known_allergies=bool(p.no_known_allergies))

@app.get("/api/patients/search", response_model=List[schemas.PatientOut])
def search_patients(q: str, db: Session = Depends(get_db)):
    res = crud.search_patients(db, q)
    return [schemas.PatientOut(id=p.id, full_name=p.full_name, date_of_birth=p.date_of_birth, sex=p.sex, no_known_allergies=bool(p.no_known_allergies)) for p in res]

@app.post("/api/patients/{patient_id}/encounters", response_model=schemas.EncounterOut)
def create_encounter(patient_id:int, payload: schemas.EncounterCreate, db: Session = Depends(get_db)):
    if not crud.get_patient(db, patient_id): raise HTTPException(404, "Patient not found")
    e = crud.create_encounter(db, patient_id, payload)
    return schemas.EncounterOut(id=e.id, patient_id=e.patient_id, encounter_datetime=e.encounter_datetime, pregnancy_status=e.pregnancy_status)

@app.get("/api/encounters/{encounter_id}/sheet", response_model=schemas.EncounterSheet)
def get_sheet(encounter_id:int, db: Session = Depends(get_db)):
    e = crud.get_encounter(db, encounter_id)
    if not e: raise HTTPException(404, "Encounter not found")
    return schemas.EncounterSheet(
        encounter=schemas.EncounterOut(id=e.id, patient_id=e.patient_id, encounter_datetime=e.encounter_datetime, pregnancy_status=e.pregnancy_status),
        items=crud.sheet_items(db, encounter_id),
    )

@app.put("/api/encounters/{encounter_id}/items/{item_type}", response_model=schemas.EncounterItemOut)
def upsert_item(encounter_id:int, item_type:str, payload: schemas.EncounterItemUpsert, db: Session = Depends(get_db)):
    allowed={"VITALS","NOTE","PMH","MEDICATION","DIAGNOSIS","PLAN","OUTCOME"}
    if item_type not in allowed: raise HTTPException(400, "Invalid item_type")
    if not crud.get_encounter(db, encounter_id): raise HTTPException(404, "Encounter not found")
    it = crud.upsert_item(db, encounter_id, item_type, payload.payload_json, payload.summary_text)
    return schemas.EncounterItemOut(id=it.id, encounter_id=it.encounter_id, item_type=it.item_type, summary_text=it.summary_text, payload_json=it.payload_json, created_at=it.created_at)

@app.get("/api/patients/{patient_id}/export.xlsx")
def export_patient_excel(patient_id: int, db: Session = Depends(get_db)):
    bundle = crud.get_patient_export_bundle(db, patient_id)
    if not bundle:
        raise HTTPException(404, "Patient not found")

    p = bundle["patient"]
    ids = bundle["identifiers"]
    encs = bundle["encounters"]
    items = bundle["items"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Patient"
    ws1.append(["Field", "Value"])
    ws1.append(["Patient ID", p.id])
    ws1.append(["Full name", p.full_name])
    ws1.append(["DOB", str(p.date_of_birth) if p.date_of_birth else ""])
    ws1.append(["Sex", p.sex])
    ws1.append(["No known allergies", bool(p.no_known_allergies)])
    # identifiers
    for pid in ids:
        ws1.append([f"Identifier ({pid.id_type})", pid.id_value])

    ws2 = wb.create_sheet("Encounters")
    ws2.append(["Encounter ID", "Patient ID", "Datetime", "Pregnancy", "Chief complaint", "Clinical summary", "Weight(kg)", "Specialty"])
    for e in encs:
        ws2.append([
            e.id, e.patient_id, str(e.encounter_datetime), e.pregnancy_status,
            e.chief_complaint or "", e.clinical_summary or "",
            float(e.weight_kg) if e.weight_kg is not None else "", e.specialty_code or ""
        ])

    ws3 = wb.create_sheet("Items")
    ws3.append(["Encounter ID", "Item type", "Summary text", "Payload JSON"])
    for e in encs:
        for it in items.get(e.id, []):
            ws3.append([e.id, it.item_type, it.summary_text or "", json.dumps(it.payload_json, ensure_ascii=False)])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"patient_{p.id}_export.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/health")
def health():
    return {"status": "ok"}
